import time

import openpyxl
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait


file_path = "C:\\Users\\Shanmugam\\Downloads\\download.xlsx"
def update_excel_data(filePath, searchTerm, colName, newValue):

    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = newValue
    book.save(filePath)


driver = webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.implicitly_wait(5)
fruit_name = "Apple"
newValue = "976"
driver.find_element(By.CSS_SELECTOR, "#downloadButton").click()

# Edit the excel with updated value
update_excel_data(file_path, fruit_name, "price", newValue)

#Upload the file
file_input = driver.find_element(By.CSS_SELECTOR, "#fileinput")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actualPrice = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
assert actualPrice == newValue